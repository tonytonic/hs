#!/usr/bin/env python3
"""
VERIFIER-CCN.PY — cohérence des sources CCN de l'application

POURQUOI CE SCRIPT ?
La même convention est décrite dans cinq fichiers indépendants. Rien ne garantit
qu'ils restent d'accord, et l'expérience montre que c'est un vrai risque :

  - en corrigeant l'IDCC 5021, il a été modifié dans deux fichiers sur cinq ;
    la divergence n'a été vue que par hasard, plusieurs étapes plus tard ;
  - l'IDCC 2120 était étiqueté « Banques populaires » dans quatre fichiers alors
    que 2120 est la convention de la banque (AFB) et que la Banque Populaire est
    le 3210. Une comparaison au référentiel officiel ne l'avait PAS détecté,
    parce que « Banques populaires » et « Banque » partagent un mot. Le
    croisement des fichiers entre eux, lui, le voit immédiatement.

Ce script fait donc DEUX contrôles complémentaires :
  A. cohérence interne  — les cinq fichiers disent-ils la même chose ?
  B. conformité externe — leurs codes existent-ils au référentiel DARES ?

Le contrôle A trouve des erreurs que B manque, et inversement. Les deux sont utiles.

USAGE
    python3 verifier-ccn.py
    python3 verifier-ccn.py --dares ccn/Dares_Suivi_Historique_convention_collective_Juin2026.xlsx
    python3 verifier-ccn.py --out rapport.md
    python3 verifier-ccn.py --strict        # code de sortie 1 si divergence

Sans --dares, seul le contrôle A est effectué (aucune dépendance à openpyxl).
"""
import argparse
import json
import os
import re
import sys
import unicodedata

# Codes internes volontaires (alias, sentinelles) : exclus des contrôles.
# Voir l'en-tête de ccn/conventions-collectives.js — la forme est « code réel x 10 »,
# pour que la recherche par startsWith() remonte les deux entrées.
ALIAS = {0, 5730, 16110, 21200, 25960, 30900, 32290}

# Motif des fichiers en notation objet.
# La capture est NON GOURMANDE et s'arrête à la première apostrophe SUIVIE D'UNE
# VIRGULE. C'est ce qui permet de traverser les apostrophes échappées à
# l'intérieur du libellé (nom: 'Cabinets d\'avocats',) : une capture naïve
# s'arrêterait sur cette apostrophe interne, tronquerait le libellé, et ferait
# croire à une divergence entre fichiers.
MOTIF_NOM = r"idcc:\s*(\d+),\s*nom:\s*'(.*?)',"

# Sources : (chemin, motif de capture (code, libellé), libellé court)
SOURCES = [
    ("GrillePaye/index.html", None, "GrillePaye/CCN_ALL"),
    ("ccn/conventions-collectives.js", r'\{i:(\d+),[^}]*?n:"([^"]+)"', "conventions-collectives"),
    ("module5/js/data/ccn-partiel.js", r'\{i:(\d+),\s*n:"([^"]+)"', "module5/ccn-partiel"),
    ("module6/ccn/conventions-cadres.js", MOTIF_NOM, "module6/conventions-cadres"),
    ("module6/js/data/ccn-adapter.js", MOTIF_NOM, "module6/ccn-adapter"),
]

# Mots trop génériques pour distinguer deux conventions.
VIDES = {"convention", "collective", "nationale", "national", "des", "les", "pour", "sur",
         "travail", "personnel", "personnels", "salaries", "entreprises", "entreprise",
         "industrie", "industries", "commerce", "commerces", "societes", "organismes",
         "etablissements", "cadres", "ouvriers", "departementale", "branche"}


def mots(t, racine=True):
    """Mots signifiants d'un libellé, normalisés.

    ARBITRAGE À CONNAÎTRE — bruit contre détection.

    Avec `racine=True` (défaut), pluriel et féminin sont gommés : « intérimaire »
    et « intérimaires » se rejoignent, ce qui évite de signaler des divergences
    qui n'en sont pas. Mais cette normalisation rapproche aussi « Banque » et
    « Banques populaires », qui sont pourtant DEUX conventions distinctes
    (IDCC 2120 et 3210) — une erreur réelle passerait alors inaperçue.

    Avec `racine=False` (option --sensible), la comparaison est littérale : elle
    remonte davantage de candidats, dont des faux positifs, mais ne laisse pas
    passer ce genre de nuance. À utiliser pour une revue périodique à la main.

    Aucun des deux réglages n'est « le bon » : le premier sert au contrôle
    automatique de non-régression, le second à l'inspection.
    """
    t = unicodedata.normalize("NFD", (t or "").lower())
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    out = set()
    for w in re.split(r"[^a-z0-9]+", t):
        if len(w) > 3 and w not in VIDES:
            if not racine:
                out.add(w)
                continue
            # Racine grossière : on retire les marques de pluriel et de féminin
            # jusqu'à stabilisation. Deux passes fixes ne suffisent pas — le
            # féminin « inadaptées » porte un « e » de plus que le masculin
            # « inadaptés », et un rabotage à profondeur fixe les laisse
            # différents. En réduisant jusqu'à racine stable, les deux donnent
            # « inadapt » et se rejoignent.
            while len(w) > 4 and w[-1] in "sxe":
                w = w[:-1]
            out.add(w)
    return out


def lire_source(racine, chemin, motif):
    p = os.path.join(racine, chemin)
    if not os.path.exists(p):
        return None
    s = open(p, encoding="utf-8", errors="replace").read()
    if motif is None:            # CCN_ALL : tableau JSON
        m = re.search(r"const CCN_ALL=(\[\[.*?\]\]);", s)
        if not m:
            return None
        # Un IDCC peut avoir plusieurs entrées (alias de recherche voulus, cf. 1979) :
        # on conserve tous les libellés pour ne pas les signaler à tort.
        d = {}
        for c in json.loads(m.group(1)):
            d.setdefault(c[0], set()).add(c[1])
        return d
    d = {}
    for a, b in re.findall(motif, s):
        d.setdefault(int(a), set()).add(b)
    return d


def charger_dares(chemin):
    try:
        import openpyxl
    except ImportError:
        print("openpyxl absent : contrôle DARES ignoré.", file=sys.stderr)
        return None
    wb = openpyxl.load_workbook(chemin, read_only=True)
    ref = {}
    for feuille in wb.sheetnames:
        rows = list(wb[feuille].iter_rows(values_only=True))
        if not rows or str(rows[0][0]).strip().upper() not in ("IDCC", "CODE"):
            continue
        for r in rows[1:]:
            if r[0] and str(r[0]).strip().isdigit():
                ref[int(str(r[0]).strip())] = str(r[1] or "")
    return ref


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--racine", default=".")
    ap.add_argument("--dares", help="Fichier DARES .xlsx (contrôle de conformité externe)")
    ap.add_argument("--out")
    ap.add_argument("--strict", action="store_true")
    ap.add_argument("--sensible", action="store_true",
                    help="Comparaison littérale des libellés : plus de candidats, "
                         "dont des faux positifs, mais ne laisse pas passer les nuances "
                         "de qualificatif (« Banque » vs « Banques populaires »). "
                         "Pour une revue à la main, pas pour un contrôle automatique.")
    args = ap.parse_args()

    sources = {}
    manquants = []
    for chemin, motif, nom in SOURCES:
        d = lire_source(args.racine, chemin, motif)
        if d is None:
            manquants.append(chemin)
        else:
            sources[nom] = d
    if not sources:
        raise SystemExit("Aucune source CCN lisible — vérifie --racine.")

    tous = set().union(*[set(d) for d in sources.values()])
    partages = [c for c in tous if sum(1 for d in sources.values() if c in d) > 1]

    divergents = []
    for c in sorted(partages):
        if c in ALIAS:
            continue
        libs = {n: d[c] for n, d in sources.items() if c in d}
        # Divergence = deux fichiers dont AUCUN libellé ne partage de mot signifiant.
        # Un fichier peut proposer plusieurs libellés (alias) : il suffit qu'un seul
        # concorde pour considérer les deux sources d'accord.
        noms = list(libs)
        conflit = False
        for i in range(len(noms)):
            for j in range(i + 1, len(noms)):
                rac = not args.sensible
                a = set().union(*[mots(x, rac) for x in libs[noms[i]]])
                b = set().union(*[mots(x, rac) for x in libs[noms[j]]])
                if a and b and not (a & b):
                    conflit = True
        if conflit:
            divergents.append((c, libs))

    ref = charger_dares(args.dares) if args.dares else None
    hors_ref = []
    if ref:
        for c in sorted(tous):
            if c in ALIAS or c in ref:
                continue
            ou = [n for n, d in sources.items() if c in d]
            hors_ref.append((c, ou))

    L = ["# Cohérence des sources CCN", ""]
    for n, d in sources.items():
        L.append(f"- `{n}` : {len(d)} code(s)")
    if manquants:
        L.append("")
        L.append("Fichiers introuvables : " + ", ".join(f"`{m}`" for m in manquants))
    L.append("")
    L.append(f"{len(tous)} codes distincts, dont **{len(partages)}** présents dans plusieurs fichiers.")
    L.append("")

    if divergents:
        L.append(f"## ⚠️ {len(divergents)} code(s) décrits différemment selon le fichier")
        L.append("")
        L.append("Un même IDCC ne peut pas désigner deux conventions. L'un des fichiers se trompe.")
        L.append("")
        for c, libs in divergents:
            L.append(f"- **IDCC {c}**")
            for n, v in libs.items():
                L.append(f"    - `{n}` : {' / '.join(sorted(v))[:70]}")
        L.append("")
    else:
        L.append("Aucune divergence entre fichiers.")
        L.append("")

    if ref is not None:
        if hors_ref:
            L.append(f"## {len(hors_ref)} code(s) absent(s) du référentiel DARES")
            L.append("")
            for c, ou in hors_ref[:50]:
                L.append(f"- IDCC {c} — présent dans : {', '.join(ou)}")
            if len(hors_ref) > 50:
                L.append(f"- … et {len(hors_ref) - 50} autre(s)")
            L.append("")
        else:
            L.append("Tous les codes existent au référentiel DARES.")
            L.append("")

    rapport = "\n".join(L)
    if args.out:
        open(args.out, "w", encoding="utf-8").write(rapport)
        print(f"Rapport écrit dans {args.out}")
    else:
        print(rapport)

    if args.strict and divergents:
        print("\nDivergence entre fichiers (--strict).", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
